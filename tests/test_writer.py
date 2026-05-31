# tests/test_writer.py
"""Tests for data/writer.py — find_row_by_com, save_unit."""

from __future__ import annotations

import pytest
from openpyxl import Workbook, load_workbook

from data.models import Unit
from data.writer import find_row_by_com, save_unit


@pytest.fixture
def workbook_with_data(tmp_path):
    """Create a real .xlsx file with COM numbers in column C."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers
    ws["A1"] = "Due Date"
    ws["B1"] = "Prev Due"
    ws["C1"] = "COM Number"
    ws["D1"] = "Mfg"
    ws["E1"] = "Detailer"
    ws["F1"] = "Job Name"
    ws["G1"] = "Contract"
    ws["H1"] = "Description"
    ws["I1"] = "Build Date"
    ws["J1"] = ""
    ws["K1"] = "Dept Hours"
    ws["L1"] = "% Complete"
    ws["N1"] = "Actual Hours"
    ws["U1"] = "Checking Status"
    ws["V1"] = "Target Hours"
    ws["W1"] = "IEC Hours"
    ws["X1"] = "Start Date"
    ws["Y1"] = "Checking Date"
    ws["Z1"] = "Completion Date"

    # Data rows
    ws["C2"] = "COM-001"
    ws["F2"] = "Job Alpha"
    ws["L2"] = 0.5
    ws["C3"] = "COM-002"
    ws["F3"] = "Job Beta"
    ws["L3"] = 0.75
    ws["C4"] = "COM-003"
    ws["F4"] = "Job Gamma"
    ws["L4"] = 1.0

    path = str(tmp_path / "test.xlsx")
    wb.save(path)
    wb.close()
    return path


# ── find_row_by_com ──────────────────────────────────────────────


class TestFindRowByCom:
    def test_finds_existing_com(self, workbook_with_data):
        wb = load_workbook(workbook_with_data)
        ws = wb["Sheet1"]
        assert find_row_by_com(ws, "COM-001") == 2
        assert find_row_by_com(ws, "COM-002") == 3
        assert find_row_by_com(ws, "COM-003") == 4
        wb.close()

    def test_returns_none_for_missing_com(self, workbook_with_data):
        wb = load_workbook(workbook_with_data)
        ws = wb["Sheet1"]
        assert find_row_by_com(ws, "COM-999") is None
        wb.close()

    def test_strips_whitespace(self, workbook_with_data):
        """Whitespace-tolerant matching."""
        wb = load_workbook(workbook_with_data)
        ws = wb["Sheet1"]
        # The function strips both the cell value and the search value
        assert find_row_by_com(ws, "  COM-001  ") == 2
        wb.close()

    def test_empty_sheet_returns_none(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["C1"] = "Header"
        path = str(tmp_path / "empty.xlsx")
        wb.save(path)
        wb.close()

        wb2 = load_workbook(path)
        assert find_row_by_com(wb2.active, "COM-001") is None
        wb2.close()


# ── save_unit ─────────────────────────────────────────────────────


class TestSaveUnit:
    def test_save_updates_fields(self, workbook_with_data):
        """Verify the function signature accepts the expected parameters."""
        import inspect

        sig = inspect.signature(save_unit)
        params = list(sig.parameters.keys())
        assert "excel_path" in params
        assert "unit" in params
        assert "sheet_name" in params

    def test_save_unit_raises_on_missing_com(self, workbook_with_data):
        unit = Unit(
            com_number="NONEXISTENT",
            job_name="X",
            contract_number="Y",
            description="Z",
            detailer="D",
            checking_status="",
        )
        with pytest.raises(ValueError, match="not found"):
            save_unit(workbook_with_data, unit, "Sheet1")

    def test_save_creates_percent_format(self, tmp_path):
        """Verify percent formatting on column L after save."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Set up columns matching COLUMN_MAP for COM and percent
        ws["C1"] = "COM Number"
        ws["C2"] = "COM-001"
        ws["L1"] = "% Complete"
        ws["L2"] = 0.0

        path = str(tmp_path / "pct_test.xlsx")
        wb.save(path)
        wb.close()

        unit = Unit(
            com_number="COM-001",
            job_name="",
            contract_number="",
            description="",
            detailer="",
            checking_status="",
            percent_complete=85.0,
        )
        save_unit(path, unit, "Sheet1")

        wb2 = load_workbook(path)
        ws2 = wb2["Sheet1"]
        cell = ws2["L2"]
        assert cell.value == 0.85
        assert cell.number_format == "0%"
        wb2.close()

    def test_save_unit_writes_correct_fields(self, tmp_path):
        """End-to-end test: write a unit and read back its values."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Set up all columns from COLUMN_MAP
        ws["C1"] = "COM Number"
        ws["C2"] = "SAVE-001"
        ws["F1"] = "Job Name"
        ws["F2"] = "Original"
        ws["E1"] = "Detailer"
        ws["E2"] = "Original Detailer"
        ws["K1"] = "Dept Hours"
        ws["K2"] = 0.0
        ws["L1"] = "% Complete"
        ws["L2"] = 0.0
        ws["U1"] = "Checking Status"
        ws["U2"] = ""
        ws["V1"] = "Target Hours"
        ws["V2"] = 0.0
        ws["W1"] = "IEC Hours"
        ws["W2"] = 0.0
        ws["A1"] = "Due Date"
        ws["A2"] = None
        ws["N1"] = "Actual Hours"
        ws["N2"] = 0.0

        path = str(tmp_path / "e2e_save.xlsx")
        wb.save(path)
        wb.close()

        unit = Unit(
            com_number="SAVE-001",
            job_name="New Job Name",
            contract_number="NEW-CONT",
            description="New desc",
            detailer="New Detailer",
            checking_status="Released",
            status_color="green",
            department_hours=100.0,
            target_department_hours=90.0,
            iec_internal_hours=10.0,
            percent_complete=50.0,
            actual_hours=25.0,
        )
        save_unit(path, unit, "Sheet1")

        wb2 = load_workbook(path)
        ws2 = wb2["Sheet1"]
        row = 2

        assert ws2[f"C{row}"].value == "SAVE-001"
        assert ws2[f"F{row}"].value == "New Job Name"
        assert ws2[f"E{row}"].value == "New Detailer"
        assert ws2[f"K{row}"].value == 100.0
        assert ws2[f"L{row}"].value == 0.5
        assert ws2[f"U{row}"].value == "Released"
        assert ws2[f"V{row}"].value == 90.0
        assert ws2[f"W{row}"].value == 10.0
        wb2.close()
