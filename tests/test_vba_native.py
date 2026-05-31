# tests/test_vba_native.py
"""Tests for automation/vba_native.py — backup, apply_formulas, coms_into_list, move_data_in."""

from __future__ import annotations

import os
from datetime import date

import pytest
from openpyxl import Workbook, load_workbook

from automation.vba_native import (
    apply_formulas,
    backup,
    coms_into_list,
    move_data_in,
    save_master,
)

# ── save_master ───────────────────────────────────────────────────


class TestSaveMaster:
    def test_is_noop(self, tmp_path):
        """save_master should do nothing (pass)."""
        path = str(tmp_path / "test.xlsx")
        # Create a minimal file so it exists
        wb = Workbook()
        wb.save(path)
        wb.close()

        # Should not raise and should do nothing
        save_master(path)
        assert os.path.exists(path)


# ── backup ─────────────────────────────────────────────────────────


class TestBackup:
    def test_creates_archive_directory(self, tmp_path):
        src = str(tmp_path / "MyWorkbook.xlsx")
        wb = Workbook()
        wb.save(src)
        wb.close()

        backup(src)
        archive_dir = tmp_path / "Archive"
        assert archive_dir.is_dir()

    def test_creates_timestamped_copy(self, tmp_path):
        src = str(tmp_path / "MyWorkbook.xlsx")
        wb = Workbook()
        wb.save(src)
        wb.close()

        backup(src)
        archive_dir = tmp_path / "Archive"
        files = list(archive_dir.glob("*.xlsm"))
        assert len(files) == 1
        # Should start with today's date
        today_str = date.today().strftime("%Y-%m-%d")
        assert files[0].name.startswith(today_str)
        assert files[0].name.endswith("_MyWorkbook.xlsm")

    def test_copies_content_correctly(self, tmp_path):
        src = str(tmp_path / "Data.xlsx")
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "test data"
        ws["B2"] = 42
        wb.save(src)
        wb.close()

        backup(src)
        archive_dir = tmp_path / "Archive"
        files = list(archive_dir.glob("*.xlsm"))
        wb2 = load_workbook(str(files[0]))
        ws2 = wb2.active
        assert ws2["A1"].value == "test data"
        assert ws2["B2"].value == 42
        wb2.close()

    def test_multiple_backups_different_names(self, tmp_path):
        """If run on different days, different archive files should be created."""
        src = str(tmp_path / "MyWorkbook.xlsx")
        wb = Workbook()
        wb.save(src)
        wb.close()

        backup(src)
        # Running again on same day should create second file with same timestamp
        backup(src)
        archive_dir = tmp_path / "Archive"
        files = list(archive_dir.glob("*.xlsm"))
        # Same timestamp means same name → shutil.copy2 overwrites
        assert len(files) == 1


# ── apply_formulas ─────────────────────────────────────────────────


class TestApplyFormulas:
    def test_formula_column_m(self):
        """Column M should get the remaining hours formula with per-row references."""
        wb = Workbook()
        ws = wb.active
        ws["A2"] = date(2025, 6, 15)
        ws["K2"] = 40.0
        ws["L2"] = 0.5

        apply_formulas(ws)

        formula = ws["M2"].value
        assert formula is not None
        assert "L2" in str(formula)
        assert "K2" in str(formula)

    def test_formula_column_m_per_row_references(self):
        """Column M formulas should use row-specific references, not hardcoded row 2."""
        wb = Workbook()
        ws = wb.active
        ws["A2"] = date(2025, 6, 15)
        ws["A3"] = date(2025, 6, 16)
        ws["K2"] = 40.0
        ws["K3"] = 50.0
        ws["L2"] = 0.5
        ws["L3"] = 0.8

        apply_formulas(ws)

        # Row 2 formula should reference L2/K2
        formula_r2 = str(ws["M2"].value)
        assert "L2" in formula_r2
        assert "K2" in formula_r2

        # Row 3 formula should reference L3/K3, NOT L2/K2
        formula_r3 = str(ws["M3"].value)
        assert "L3" in formula_r3
        assert "K3" in formula_r3
        assert "L2" not in formula_r3
        assert "K2" not in formula_r3

    def test_formula_column_r_per_row_references(self):
        """Column R SUMIFS should use row-specific references and proper A1 range notation."""
        wb = Workbook()
        ws = wb.active
        ws["A2"] = date(2025, 6, 15)
        ws["A3"] = date(2025, 6, 16)
        ws["O2"] = date(2025, 6, 20)
        ws["O3"] = date(2025, 6, 27)

        apply_formulas(ws)

        # Row 2 formula should reference O2
        formula_r2 = str(ws["R2"].value)
        assert "O2" in formula_r2
        assert "SUMIFS(K$" in formula_r2  # proper range, not C11

        # Row 3 formula should reference O3, not O2
        formula_r3 = str(ws["R3"].value)
        assert "O3" in formula_r3

    def test_formula_column_t(self):
        """Column T should get the percentage formula."""
        wb = Workbook()
        ws = wb.active
        ws["A2"] = date(2025, 6, 15)

        apply_formulas(ws)

        formula = ws["T2"].value
        assert formula is not None

    def test_number_format_column_l(self):
        """Column L should be formatted as percentage."""
        wb = Workbook()
        ws = wb.active
        ws["A2"] = date(2025, 6, 15)
        ws["L2"] = 0.5

        apply_formulas(ws)

        assert ws["L2"].number_format == "0%"

    def test_number_format_column_t(self):
        """Column T should be formatted as percentage."""
        wb = Workbook()
        ws = wb.active
        ws["A2"] = date(2025, 6, 15)

        apply_formulas(ws)

        assert ws["T2"].number_format == "0%"

    def test_skips_empty_rows(self):
        """Rows with empty column A should be skipped."""
        wb = Workbook()
        ws = wb.active
        ws["A2"] = None
        ws["A3"] = date(2025, 6, 15)

        apply_formulas(ws)

        assert ws["M2"].value is None  # skipped
        assert ws["M3"].value is not None  # processed

    def test_skips_empty_string_rows(self):
        """Rows with empty string column A should be skipped."""
        wb = Workbook()
        ws = wb.active
        ws["A2"] = "   "

        apply_formulas(ws)

        assert ws["M2"].value is None

    def test_processes_multiple_rows(self):
        """Multiple data rows should all get formulas."""
        wb = Workbook()
        ws = wb.active
        for i in range(2, 6):
            ws[f"A{i}"] = date(2025, 6, i)

        apply_formulas(ws)

        for i in range(2, 6):
            assert ws[f"M{i}"].value is not None
            assert ws[f"T{i}"].value is not None


# ── coms_into_list ────────────────────────────────────────────────


class TestComsIntoList:
    def _make_target_workbook(self, tmp_path, unedited_rows=None, existing_rows=None):
        """Create a target workbook with Unedited Report + Current List sheets."""
        wb = Workbook()

        # Current List sheet
        ws_current = wb.active
        ws_current.title = "Current List"
        ws_current["A1"] = "New Due"
        ws_current["B1"] = "Old Due"
        ws_current["C1"] = "COM"
        ws_current["D1"] = "Mfg"
        ws_current["E1"] = "Mfg2"
        ws_current["F1"] = "Job"
        ws_current["G1"] = "Contract"
        ws_current["H1"] = "Desc"
        ws_current["I1"] = "Build"
        ws_current["J1"] = "Cycle"
        ws_current["K1"] = "Hrs"

        if existing_rows:
            for idx, row in enumerate(existing_rows, start=2):
                for col_idx, val in enumerate(row, start=1):
                    ws_current.cell(row=idx, column=col_idx, value=val)

        # Unedited Report sheet
        ws_unedited = wb.create_sheet("Unedited Report")
        ws_unedited["A1"] = "Date"
        ws_unedited["B1"] = "COM"
        ws_unedited["C1"] = "Mfg"
        ws_unedited["D1"] = "Job"
        ws_unedited["E1"] = "Contract"
        ws_unedited["F1"] = "Desc"
        ws_unedited["G1"] = "Build"
        ws_unedited["H1"] = "Cycle"
        ws_unedited["I1"] = "Hrs"

        if unedited_rows:
            for idx, row in enumerate(unedited_rows, start=2):
                for col_idx, val in enumerate(row, start=1):
                    ws_unedited.cell(row=idx, column=col_idx, value=val)

        path = str(tmp_path / "target.xlsx")
        wb.save(path)
        wb.close()
        return path

    def test_inserts_new_row(self, tmp_path):
        """A COM that doesn't exist should be appended as a new row."""
        unedited = [
            [
                date(2025, 6, 20),
                "COM-NEW",
                "MFG1",
                "JOB1",
                "C1",
                "D1",
                date(2025, 7, 1),
                "CYC1",
                40.0,
            ],
        ]
        path = self._make_target_workbook(tmp_path, unedited_rows=unedited)
        rows = coms_into_list(path)
        assert rows == 1

        wb = load_workbook(path)
        ws = wb["Current List"]
        # Row 2 should now have COM-NEW
        com_col = 3  # C
        found = False
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=com_col).value == "COM-NEW":
                found = True
                break
        assert found, "COM-NEW was not inserted into Current List"
        wb.close()

    def test_updates_existing_row_date(self, tmp_path):
        """An existing COM with a shifted date should update columns."""
        unedited = [
            [
                date(2025, 6, 20),
                "COM-OLD",
                "MFG1",
                "JOB1",
                "C1",
                "D1",
                date(2025, 7, 1),
                "CYC1",
                40.0,
            ],
        ]
        existing = [
            [date(2025, 6, 10), None, "COM-OLD"],  # Date in A, COM in C
        ]
        path = self._make_target_workbook(tmp_path, unedited_rows=unedited, existing_rows=existing)
        rows = coms_into_list(path)
        assert rows == 1

        wb = load_workbook(path)
        ws = wb["Current List"]
        # Row 2: A should be new date, B should be old date
        # openpyxl may return datetime for date cells
        a_val = ws["A2"].value
        b_val = ws["B2"].value
        if hasattr(a_val, "date"):
            a_val = a_val.date()
        if hasattr(b_val, "date"):
            b_val = b_val.date()
        assert a_val == date(2025, 6, 20)
        assert b_val == date(2025, 6, 10)
        wb.close()

    def test_returns_row_count(self, tmp_path):
        """Should return the number of rows processed from Unedited Report."""
        unedited = [
            [date(2025, 6, 20), "COM-A", "MFG", "JOB", "C", "D", date(2025, 7, 1), "CYC", 40.0],
            [date(2025, 6, 21), "COM-B", "MFG", "JOB", "C", "D", date(2025, 7, 2), "CYC", 50.0],
            [date(2025, 6, 22), "COM-C", "MFG", "JOB", "C", "D", date(2025, 7, 3), "CYC", 60.0],
        ]
        path = self._make_target_workbook(tmp_path, unedited_rows=unedited)
        rows = coms_into_list(path)
        assert rows == 3

    def test_skips_empty_com_rows(self, tmp_path):
        """Rows with empty COM column in Unedited Report should be skipped."""
        unedited = [
            [date(2025, 6, 20), None, "MFG", "JOB", "C", "D", date(2025, 7, 1), "CYC", 40.0],
            [date(2025, 6, 21), "COM-X", "MFG", "JOB", "C", "D", date(2025, 7, 2), "CYC", 50.0],
            [date(2025, 6, 22), "", "MFG", "JOB", "C", "D", date(2025, 7, 3), "CYC", 60.0],
        ]
        path = self._make_target_workbook(tmp_path, unedited_rows=unedited)
        rows = coms_into_list(path)
        assert rows == 1

    def test_applies_formulas_after_merge(self, tmp_path):
        """After merging, formulas should be applied to Current List."""
        unedited = [
            [date(2025, 6, 20), "COM-F", "MFG", "JOB", "C", "D", date(2025, 7, 1), "CYC", 40.0],
        ]
        path = self._make_target_workbook(tmp_path, unedited_rows=unedited)
        coms_into_list(path)

        wb = load_workbook(path)
        ws = wb["Current List"]
        assert ws["M2"].value is not None  # formula applied
        assert ws["T2"].value is not None
        wb.close()


# ── move_data_in ──────────────────────────────────────────────────


class TestMoveDataIn:
    def _make_source_workbook(self, tmp_path, data_rows):
        """Create a source workbook with a SCHDetailingReport sheet."""
        wb = Workbook()
        ws = wb.active
        ws.title = "SCHDetailingReport"
        # Header
        headers = [
            "Date",
            "COM",
            "Mfg",
            "Job",
            "Contract",
            "Desc",
            "Build",
            "Cycle",
            "DeptHrs",
            "ColJ",
            "ColK",
            "WeekEnding",
        ]
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=h)

        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        path = str(tmp_path / "source.xlsx")
        wb.save(path)
        wb.close()
        return path

    def test_full_pipeline(self, tmp_path):
        """Test move_data_in end-to-end: read → write → merge → backup."""
        data = [
            [
                date(2025, 7, 1),
                "COM-P",
                "MFG",
                "JOB",
                "CON",
                "D",
                date(2025, 8, 1),
                "C1",
                30.0,
                None,
                None,
                "07/01/2025",
            ],
        ]
        source = self._make_source_workbook(tmp_path, data)

        # Create target workbook with both sheets
        wb = Workbook()
        ws_unedited = wb.active
        ws_unedited.title = "Unedited Report"
        for i in range(1, 15):
            ws_unedited.cell(row=1, column=i, value=f"H{i}")
        ws_current = wb.create_sheet("Current List")
        for i in range(1, 15):
            ws_current.cell(row=1, column=i, value=f"H{i}")
        target = str(tmp_path / "target.xlsx")
        wb.save(target)
        wb.close()

        rows = move_data_in(source, target)
        assert rows == 1

        # Verify Unedited Report has data
        wb2 = load_workbook(target)
        ws_u = wb2["Unedited Report"]
        assert ws_u["B2"].value == "COM-P"
        wb2.close()

        # Verify backup was created
        archive_dir = tmp_path / "Archive"
        assert archive_dir.is_dir()
        files = list(archive_dir.glob("*.xlsm"))
        assert len(files) >= 1

    def test_raises_on_empty_source(self, tmp_path):
        """Should raise ValueError if SCHDetailingReport has no data rows."""
        source = self._make_source_workbook(tmp_path, [])

        wb = Workbook()
        ws = wb.active
        ws.title = "Unedited Report"
        for i in range(1, 15):
            ws.cell(row=1, column=i, value=f"H{i}")
        ws2 = wb.create_sheet("Current List")
        for i in range(1, 15):
            ws2.cell(row=1, column=i, value=f"H{i}")
        target = str(tmp_path / "target.xlsx")
        wb.save(target)
        wb.close()

        with pytest.raises(ValueError, match="No data found"):
            move_data_in(source, target)

    def test_raises_on_missing_sheet(self, tmp_path):
        """Should handle the case where source doesn't have SCHDetailingReport."""
        wb = Workbook()
        wb.active.title = "WrongSheet"
        source = str(tmp_path / "bad_source.xlsx")
        wb.save(source)
        wb.close()

        wb2 = Workbook()
        ws = wb2.active
        ws.title = "Unedited Report"
        wb2.create_sheet("Current List")
        target = str(tmp_path / "target.xlsx")
        wb2.save(target)
        wb2.close()

        with pytest.raises(KeyError):
            move_data_in(source, target)
