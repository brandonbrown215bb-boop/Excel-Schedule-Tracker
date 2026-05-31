# tests/conftest.py
"""Shared fixtures for the test suite."""

from __future__ import annotations

import tempfile
from datetime import date, timedelta
from pathlib import Path

import pytest
from openpyxl import Workbook

from data.models import Unit

# ── Unit fixtures ────────────────────────────────────────────────


@pytest.fixture
def sample_unit() -> Unit:
    """A basic Unit with typical values."""
    return Unit(
        com_number="COM-001",
        job_name="Test Job Alpha",
        contract_number="CNT-2024-001",
        description="A sample unit for testing",
        detailer="Jackie / IEC Internals",
        checking_status="Not Started",
        status_color="gray",
        department_hours=40.0,
        target_department_hours=36.0,
        iec_internal_hours=4.0,
        percent_complete=50.0,
        actual_hours=20.0,
        working_days=[0, 1, 2, 3],
        unit_detailing_start_date=date(2025, 1, 15),
        unit_moved_to_checking_date=None,
        unit_detailing_completion_date=None,
        dept_due_date_previous=date(2025, 3, 1),
        detailing_due_date=date.today() + timedelta(days=30),
        build_date=date(2025, 6, 1),
    )


@pytest.fixture
def unassigned_unit() -> Unit:
    """A Unit at 0% with no due date — should be gray."""
    return Unit(
        com_number="COM-002",
        job_name="Unassigned Job",
        contract_number="CNT-2024-002",
        description="No detailer assigned yet",
        detailer="— Unassigned —",
        checking_status="",
        status_color="gray",
        department_hours=20.0,
        percent_complete=0.0,
        working_days=[0, 1, 2, 3],
        detailing_due_date=None,
        build_date=None,
    )


@pytest.fixture
def completed_unit() -> Unit:
    """A Unit at 100% — should always be green."""
    return Unit(
        com_number="COM-003",
        job_name="Finished Job",
        contract_number="CNT-2024-003",
        description="All done",
        detailer="Maria / RGV Team",
        checking_status="Released",
        status_color="green",
        department_hours=80.0,
        percent_complete=100.0,
        working_days=[0, 1, 2, 3],
        detailing_due_date=date.today() - timedelta(days=10),
        build_date=date.today() - timedelta(days=5),
    )


@pytest.fixture
def overdue_unit() -> Unit:
    """A Unit past its due date — capacity check will flag red."""
    return Unit(
        com_number="COM-004",
        job_name="Overdue Job",
        contract_number="CNT-2024-004",
        description="Way behind schedule",
        detailer="Chen / HOU Team",
        checking_status="In Progress",
        status_color="red",
        department_hours=500.0,
        percent_complete=10.0,
        working_days=[1, 2, 3, 4],
        detailing_due_date=date.today() - timedelta(days=5),
        build_date=None,
    )


@pytest.fixture
def unit_list(sample_unit, unassigned_unit, completed_unit, overdue_unit) -> list[Unit]:
    """A list of units covering all status colors."""
    return [sample_unit, unassigned_unit, completed_unit, overdue_unit]


# ── Mock workbook fixtures ───────────────────────────────────────


@pytest.fixture
def mock_workbook():
    """Create a minimal openpyxl Workbook with a 'Current List' sheet and headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Current List"

    # Row 1: headers (matching the COLUMN_MAP layout)
    headers = [
        "Detailing Due Date",
        "Prev Due Date",
        "COM Number",
        "",
        "Detailer",
        "Job Name",
        "Contract",
        "Description",
        "Build Date",
        "",
        "Dept Hours",
        "% Complete",
        "Remaining Hours",
        "Actual Hours",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "Checking Status",
        "Target Hours",
        "IEC Hours",
        "",
        "",
        "",
        "",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    return wb


@pytest.fixture
def mock_workbook_with_units(mock_workbook, unit_list):
    """A mock workbook pre-populated with unit data."""
    ws = mock_workbook["Current List"]

    from data.loader import COLUMN_MAP

    for row_idx, unit in enumerate(unit_list, start=2):
        for field_name, col_letter in COLUMN_MAP.items():
            col = ord(col_letter.upper()) - ord("A") + 1
            value = getattr(unit, field_name, None)
            if value is not None:
                if isinstance(value, date):
                    ws.cell(row=row_idx, column=col, value=value)
                elif field_name == "percent_complete":
                    ws.cell(row=row_idx, column=col, value=value / 100.0)
                else:
                    ws.cell(row=row_idx, column=col, value=value)

    return mock_workbook


@pytest.fixture
def temp_excel_file(tmp_path):
    """Create a temporary .xlsx file path for testing writer functions."""
    return str(tmp_path / "test_workbook.xlsx")


@pytest.fixture
def temp_dir():
    """Provide a temporary directory that is cleaned up after the test."""
    with tempfile.TemporaryDirectory() as d:
        yield Path(d)
