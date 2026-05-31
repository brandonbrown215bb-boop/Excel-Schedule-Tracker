# automation/vba_native.py
"""
Python implementations of the VBA macros from the Excel workbook.

Replaces the need for COM/win32com to run these macros, enabling
cross-platform operation and eliminating the keep_vba=True requirement
when saving the workbook.
"""

from __future__ import annotations

import os
import shutil
from datetime import datetime
from typing import Any

from openpyxl import load_workbook  # type: ignore[reportMissingImports]

# ── Save / Backup (replaces Save_Backup.bas) ───────────────────


def save_master(target_path: str) -> None:
    """Save the workbook to the master path (same as VBA Sub Save)."""
    # The workbook is already at target_path; just ensure it's saved.
    # openpyxl handles this when calling wb.save() — this is a no-op
    # placeholder for API parity with the VBA macro.
    pass


def backup(target_path: str) -> None:
    """Save a timestamped archive copy (replaces VBA Sub Backup)."""
    archive_dir = os.path.join(os.path.dirname(target_path), "Archive")
    os.makedirs(archive_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d")
    base = os.path.splitext(os.path.basename(target_path))[0]
    dest = os.path.join(archive_dir, f"{ts}_{base}.xlsm")
    shutil.copy2(target_path, dest)


# ── COMs_into_List (replaces Move_Data_to_Current.bas) ─────────


def coms_into_list(target_path: str) -> int:
    """
    Read rows from 'Unedited Report' sheet and merge them into
    'Current List' sheet — pure-Python equivalent of the VBA
    COMs_into_List macro.

    Returns the number of rows processed.
    """
    UNEDITED_SHEET = "Unedited Report"
    CURRENT_SHEET = "Current List"
    wb = load_workbook(target_path, data_only=True, keep_vba=False)
    try:
        ws_unedited = wb[UNEDITED_SHEET]
        ws_current = wb[CURRENT_SHEET]

        # Build a lookup of existing COM numbers in Current List (column C)
        com_col = 3  # column C
        existing: dict[str, int] = {}  # com_number → row index
        for row_idx in range(2, ws_current.max_row + 1):
            val = ws_current.cell(row=row_idx, column=com_col).value
            if val is not None:
                existing[str(val).strip()] = row_idx

        # Find first empty row in Current List for new entries
        next_empty_row = ws_current.max_row + 1

        rows_processed = 0
        max_unedited = ws_unedited.max_row or 1
        for i in range(2, max_unedited + 1):
            new_com = ws_unedited.cell(row=i, column=2).value  # col B = COM
            if new_com is None or str(new_com).strip() == "":
                continue
            new_com = str(new_com).strip()
            new_date = ws_unedited.cell(row=i, column=1).value  # col A = date
            mfg_location = ws_unedited.cell(row=i, column=3).value  # col C
            job_name = ws_unedited.cell(row=i, column=4).value  # col D
            contract = ws_unedited.cell(row=i, column=5).value  # col E
            desc = ws_unedited.cell(row=i, column=6).value  # col F
            build_date = ws_unedited.cell(row=i, column=7).value  # col G
            build_cycle = ws_unedited.cell(row=i, column=8).value  # col H
            dept_hrs = ws_unedited.cell(row=i, column=9).value  # col I
            week_ending_raw = ws_unedited.cell(row=i, column=12).value  # col L
            week_ending = str(week_ending_raw)[-10:] if week_ending_raw else ""

            if new_com in existing:
                # Update existing row if date changed
                row = existing[new_com]
                current_date = ws_current.cell(row=row, column=1).value
                if str(current_date) != str(new_date):
                    ws_current.cell(
                        row=row, column=2
                    ).value = current_date  # move old date to col B
                    ws_current.cell(row=row, column=1).value = new_date  # new date to col A
                    ws_current.cell(
                        row=row, column=4
                    ).value = mfg_location  # col E (offset +1 from C)
                    ws_current.cell(row=row, column=6).value = job_name  # col F
                    ws_current.cell(row=row, column=7).value = contract  # col G
                    ws_current.cell(row=row, column=8).value = desc  # col H
                    ws_current.cell(row=row, column=9).value = build_date  # col I
                    ws_current.cell(row=row, column=10).value = build_cycle  # col J
                    ws_current.cell(row=row, column=11).value = dept_hrs  # col K
                    ws_current.cell(row=row, column=14).value = week_ending  # col N
            else:
                # Insert new row
                row = next_empty_row
                next_empty_row += 1
                ws_current.cell(row=row, column=com_col).value = new_com
                ws_current.cell(row=row, column=1).value = new_date
                ws_current.cell(row=row, column=4).value = mfg_location
                ws_current.cell(row=row, column=6).value = job_name
                ws_current.cell(row=row, column=7).value = contract
                ws_current.cell(row=row, column=8).value = desc
                ws_current.cell(row=row, column=9).value = build_date
                ws_current.cell(row=row, column=10).value = build_cycle
                ws_current.cell(row=row, column=11).value = dept_hrs
                ws_current.cell(row=row, column=14).value = week_ending
                existing[new_com] = row

            rows_processed += 1

        # Apply formulas
        apply_formulas(ws_current)

        wb.save(target_path)
        return rows_processed
    finally:
        wb.close()


# ── ApplyFormulas (replaces Formulas.bas) ───────────────────────


def apply_formulas(ws: Any) -> None:
    """
    Apply the same formulas the VBA ApplyFormulas macro sets.
    Operates on an openpyxl worksheet in-place.

    Column mapping:
        K (11) = department_hours
        L (12) = percent_complete (input, e.g. 0.50)
        M (13) = remaining_hours (computed: dept_hrs × (1 - pct))
        N (14) = actual_hours
        O (15) = week_ending_date
        R (18) = weekly_dept_hours_sum
        S (19) = weekly_actual_hours_sum
        T (20) = percentage (computed: (R - S) / R)
    """
    max_row = ws.max_row or 1
    for i in range(2, max_row + 1):
        a_val = ws.cell(row=i, column=1).value
        if a_val is None or str(a_val).strip() == "":
            continue

        # Column M (13): Remaining Hours
        ws.cell(row=i, column=13).value = (
            f'=IF(OR(L{i}="",L{i}=0),K{i},K{i}-K{i}*L{i})'
        )
        # Column R (18): Dept hours sum for the week ending in column O
        ws.cell(row=i, column=18).value = (
            f'=SUMIFS(K$2:K${max_row},O$2:O${max_row},">="&'
            f'IF(MONTH(O{i}-WEEKDAY(O{i}-2))<>MONTH(O{i}),'
            f'EOMONTH(O{i},-1)+1,O{i}-WEEKDAY(O{i}-2)),'
            f'O$2:O${max_row},"<="&O{i})'
        )
        # Column S (19): Actual hours sum for the week ending in column O
        ws.cell(row=i, column=19).value = (
            f'=SUMIFS(N$2:N${max_row},O$2:O${max_row},">="&'
            f'IF(MONTH(O{i}-WEEKDAY(O{i}-2))<>MONTH(O{i}),'
            f'EOMONTH(O{i},-1)+1,O{i}-WEEKDAY(O{i}-2)),'
            f'O$2:O${max_row},"<="&O{i})'
        )
        # Column T (20): Percentage (compare this row vs next row week-ending)
        ws.cell(row=i, column=20).value = (
            f'=IF(R{i}="","",IF(O{i}=O{i+1},"",'
            f'(R{i}-S{i})/R{i}))'
        )

    # Number formats
    for col_letter in ("L", "R", "S"):
        for row in range(2, max_row + 1):
            ws[f"{col_letter}{row}"].number_format = "0.00"
    for row in range(2, max_row + 1):
        ws[f"L{row}"].number_format = "0%"
        ws[f"T{row}"].number_format = "0%"
    for row in range(2, max_row + 1):
        ws[f"M{row}"].number_format = "0.00"


# ── Move_Data_In (replaces Move_Unedited_In.bas) ───────────────


def move_data_in(source_path: str, target_path: str) -> int:
    """
    Import data from an unedited report file into the target workbook's
    'Unedited Report' sheet, then run coms_into_list + backup + save.

    Returns the number of rows imported.
    """
    SOURCE_SHEET = "SCHDetailingReport"
    TARGET_SHEET = "Unedited Report"

    # Read source data
    wb_src = load_workbook(source_path, read_only=True, data_only=True)
    try:
        ws_src = wb_src[SOURCE_SHEET]
        rows = []
        for row_idx, row in enumerate(
            ws_src.iter_rows(min_col=1, max_col=14, values_only=True), start=1
        ):
            if row_idx == 1:
                continue
            rows.append(list(row))
    finally:
        wb_src.close()

    if not rows:
        raise ValueError(f"No data found in '{SOURCE_SHEET}' sheet of {source_path}")

    # Write to Unedited Report sheet
    wb_tgt = load_workbook(target_path, keep_vba=False)
    try:
        ws_tgt = wb_tgt[TARGET_SHEET]

        # Clear old data (keep header row 1)
        for row in ws_tgt.iter_rows(min_row=2, max_row=ws_tgt.max_row, min_col=1, max_col=14):
            for cell in row:
                cell.value = None

        # Write new data
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_tgt.cell(row=row_idx, column=col_idx, value=value)
                if col_idx in (1, 7) and value is not None:
                    cell.number_format = "mm/dd/yy;@"

        wb_tgt.save(target_path)
    finally:
        wb_tgt.close()

    # Now run the merge (COMs_into_List) on the target
    coms_into_list(target_path)

    # Backup
    backup(target_path)

    return len(rows)
