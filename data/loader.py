# data/loader.py
import csv
import hashlib
import json
import logging
import os
import pickle
import time
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from zipfile import BadZipFile

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook  # type: ignore[reportAssignmentType]
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:  # pragma: no cover

    def load_workbook(*args, **kwargs):
        """Fallback stub when openpyxl is not installed."""
        raise ImportError(
            "openpyxl is required to load Excel files. Please install it via "
            "`pip install openpyxl`."
        )

    class InvalidFileException(Exception):  # noqa: N818 - mirrors openpyxl's exception name
        pass

from .models import Unit

CACHE_SCHEMA_VERSION = 2

# data/loader.py — corrected column map
COLUMN_MAP = {
    "detailing_due_date": "A",  # col A — Detailing Due Date
    "dept_due_date_previous": "B",  # col B — Prev Due Date
    "com_number": "C",  # col C — COM Number
    "detailer": "E",  # col E — Detailer
    "job_name": "F",  # col F — Job Name
    "contract_number": "G",  # col G — Contract
    "description": "H",  # col H — Description
    "build_date": "I",  # col I — Build Date
    "department_hours": "K",  # col K — Dept Hours
    "percent_complete": "L",  # col L = Past %
    # col M = Remaining Hours (formula)
    "actual_hours": "N",  # col N — Actual Hours
    # col O = Week Ending — not in your original columns
    "checking_status": "U",  # col U — Checking Status
    "target_department_hours": "V",  # col V — Target Dept Hours
    "iec_internal_hours": "W",  # col W — IEC Internal Hours
    "unit_detailing_start_date": "X",
    "unit_moved_to_checking_date": "Y",
    "unit_detailing_completion_date": "Z",
}

# CSV cache: all field names we persist, in a stable order.
_CSV_FIELDS = [
    "com_number",
    "job_name",
    "contract_number",
    "description",
    "detailer",
    "checking_status",
    "department_hours",
    "actual_hours",
    "target_department_hours",
    "iec_internal_hours",
    "percent_complete",
    "unit_detailing_start_date",
    "unit_moved_to_checking_date",
    "unit_detailing_completion_date",
    "dept_due_date_previous",
    "detailing_due_date",
    "build_date",
]


@dataclass
class WorkbookCache:
    """Pickle cache payload for cache-first app loads."""

    units: list[Unit]
    row_by_com: dict[str, int] = field(default_factory=dict)
    fingerprint_by_com: dict[str, str] = field(default_factory=dict)
    excel_mtime_ns: int = 0
    excel_size: int = 0
    schema_version: int = CACHE_SCHEMA_VERSION

    @classmethod
    def from_pickle(cls, cached: object) -> "WorkbookCache":
        """Read current or legacy pickle cache data."""
        if isinstance(cached, cls):
            return cached
        if isinstance(cached, dict) and "units" in cached:
            return cls(
                units=cached.get("units", []),
                row_by_com=cached.get("row_by_com", {}),
                fingerprint_by_com=cached.get("fingerprint_by_com", {}),
                excel_mtime_ns=cached.get("excel_mtime_ns", 0),
                excel_size=cached.get("excel_size", 0),
                schema_version=cached.get("schema_version", 1),
            )
        if isinstance(cached, list):
            return cls(units=cached)
        return cls(units=[])


def _cache_path(excel_path: str) -> str:
    """Return the path for the Pickle cache next to the Excel file."""
    base = os.path.splitext(excel_path)[0]
    return base + "_cache.pkl"


def _csv_cache_path(excel_path: str) -> str:
    """Return the path for the CSV cache next to the Excel file (legacy)."""
    base = os.path.splitext(excel_path)[0]
    return base + "_cache.csv"


def _workbook_signature(excel_path: str) -> tuple[int, int]:
    """Return a cheap signature for detecting workbook changes."""
    try:
        stat = os.stat(excel_path)
        return stat.st_mtime_ns, stat.st_size
    except OSError:
        return 0, 0


def parse_date(cell_value: object) -> date | None:
    if cell_value is None:
        return None
    if isinstance(cell_value, datetime):
        return cell_value.date()
    if isinstance(cell_value, date):
        return cell_value
    if isinstance(cell_value, str):
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%m-%d-%y"):
            try:
                return datetime.strptime(cell_value.strip(), fmt).date()
            except ValueError:
                continue
        logger.warning("Unrecognized date string: %r", cell_value)
    # Excel serial date (number of days since 1900-01-00)
    if isinstance(cell_value, (int, float)) and 1 < cell_value < 3000000:
        try:
            from datetime import datetime as _dt

            return _dt(1899, 12, 30) + timedelta(days=int(cell_value))
        except (ValueError, OverflowError):
            pass
    return None


def _date_to_str(d: date | None) -> str:
    """Serialize a date for the CSV cache."""
    if d is None:
        return ""
    return d.isoformat()


def parse_float(cell_value: object) -> float:
    if cell_value is None:
        return 0.0
    if isinstance(cell_value, (int, float)):
        return float(cell_value)
    if isinstance(cell_value, str):
        try:
            return float(cell_value.strip())
        except (ValueError, TypeError):
            return 0.0
    return 0.0


def unit_fingerprint(unit: Unit) -> str:
    """Stable hash of editable unit fields for optimistic conflict checks."""
    payload = {
        "com_number": unit.com_number,
        "job_name": unit.job_name,
        "contract_number": unit.contract_number,
        "description": unit.description,
        "detailer": unit.detailer,
        "checking_status": unit.checking_status,
        "department_hours": unit.department_hours,
        "actual_hours": unit.actual_hours,
        "target_department_hours": unit.target_department_hours,
        "iec_internal_hours": unit.iec_internal_hours,
        "percent_complete": unit.percent_complete,
        "unit_detailing_start_date": _date_to_str(unit.unit_detailing_start_date),
        "unit_moved_to_checking_date": _date_to_str(unit.unit_moved_to_checking_date),
        "unit_detailing_completion_date": _date_to_str(unit.unit_detailing_completion_date),
        "dept_due_date_previous": _date_to_str(unit.dept_due_date_previous),
        "detailing_due_date": _date_to_str(unit.detailing_due_date),
        "build_date": _date_to_str(unit.build_date),
    }
    raw = json.dumps(payload, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


# UnitData and UnitRowMapper removed — were empty classes with no methods or consumers.


# ── CSV helpers ──────────────────────────────────────────────────


def _unit_to_csv_row(unit: Unit) -> dict[str, str]:
    """Convert a Unit to a dict suitable for csv.DictWriter."""
    return {
        "com_number": unit.com_number,
        "job_name": unit.job_name,
        "contract_number": unit.contract_number,
        "description": unit.description,
        "detailer": unit.detailer,
        "checking_status": unit.checking_status,
        "department_hours": str(unit.department_hours),
        "actual_hours": str(unit.actual_hours),
        "target_department_hours": str(unit.target_department_hours),
        "iec_internal_hours": str(unit.iec_internal_hours),
        "percent_complete": str(unit.percent_complete),
        "unit_detailing_start_date": _date_to_str(unit.unit_detailing_start_date),
        "unit_moved_to_checking_date": _date_to_str(unit.unit_moved_to_checking_date),
        "unit_detailing_completion_date": _date_to_str(unit.unit_detailing_completion_date),
        "dept_due_date_previous": _date_to_str(unit.dept_due_date_previous),
        "detailing_due_date": _date_to_str(unit.detailing_due_date),
        "build_date": _date_to_str(unit.build_date),
    }


def _load_units_from_csv(csv_path: str, detailer_schedules: dict) -> list[Unit]:
    """Load Unit objects from the CSV cache — fast."""
    default_schedule = detailer_schedules.get("default", [0, 1, 2, 3])
    units: list[Unit] = []
    today = datetime.now().date()
    three_months_back = today - timedelta(days=90)
    twelve_months_forward = today + timedelta(days=365)

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            detailer_name = row.get("detailer", "")
            schedule = detailer_schedules.get(detailer_name, default_schedule)
            dates = [
                parse_date(row.get(fd))
                for fd in (
                    "unit_detailing_start_date",
                    "unit_moved_to_checking_date",
                    "unit_detailing_completion_date",
                    "dept_due_date_previous",
                    "detailing_due_date",
                    "build_date",
                )
            ]

            # Date filter — inline to avoid creating Unit prematurely.
            should_include = any(
                d and three_months_back <= d <= twelve_months_forward for d in dates
            )
            if not should_include:
                continue

            unit = Unit(
                com_number=row.get("com_number", ""),
                job_name=row.get("job_name", ""),
                contract_number=row.get("contract_number", ""),
                description=row.get("description", ""),
                detailer=detailer_name,
                checking_status=row.get("checking_status", ""),
                department_hours=parse_float(row.get("department_hours")),
                working_days=schedule,
                target_department_hours=parse_float(row.get("target_department_hours")),
                iec_internal_hours=parse_float(row.get("iec_internal_hours")),
                percent_complete=parse_float(row.get("percent_complete")),
                unit_detailing_start_date=dates[0],
                unit_moved_to_checking_date=dates[1],
                unit_detailing_completion_date=dates[2],
                dept_due_date_previous=dates[3],
                detailing_due_date=dates[4],
                build_date=dates[5],
            )
            unit.status_color = unit.calculated_status_color
            units.append(unit)

    return units


def _load_units_from_pickle(pickle_path: str, detailer_schedules: dict) -> list[Unit]:
    """Load Unit objects from the Pickle cache — very fast, no parsing needed."""
    default_schedule = detailer_schedules.get("default", [0, 1, 2, 3])
    today = datetime.now().date()
    three_months_back = today - timedelta(days=90)
    twelve_months_forward = today + timedelta(days=365)

    with open(pickle_path, "rb") as f:
        payload = WorkbookCache.from_pickle(pickle.load(f))

    units: list[Unit] = []
    for unit in payload.units:
        # Apply date filter
        should_include = any(
            d and three_months_back <= d <= twelve_months_forward
            for d in (
                unit.unit_detailing_start_date,
                unit.unit_moved_to_checking_date,
                unit.unit_detailing_completion_date,
                unit.dept_due_date_previous,
                unit.detailing_due_date,
                unit.build_date,
            )
        )
        if not should_include:
            continue

        # Update schedule from current config
        unit.working_days = detailer_schedules.get(unit.detailer, default_schedule)
        unit.status_color = unit.calculated_status_color
        if unit.com_number in payload.row_by_com:
            unit.excel_row = payload.row_by_com[unit.com_number]
        if unit.com_number in payload.fingerprint_by_com:
            unit.fingerprint = payload.fingerprint_by_com[unit.com_number]
        units.append(unit)

    return units


def _save_pickle_cache(excel_path: str, units: list[Unit]) -> None:
    """Write the current unit list to Pickle cache — fast binary format."""
    pickle_path = _cache_path(excel_path)
    should_wrap = all(isinstance(unit, Unit) for unit in units)
    payload: object
    if should_wrap:
        mtime_ns, size = _workbook_signature(excel_path)
        row_by_com = {
            unit.com_number: row
            for unit in units
            if (row := unit.excel_row) is not None
        }
        fingerprint_by_com = {unit.com_number: unit_fingerprint(unit) for unit in units}
        payload = WorkbookCache(
            units=units,
            row_by_com=row_by_com,
            fingerprint_by_com=fingerprint_by_com,
            excel_mtime_ns=mtime_ns,
            excel_size=size,
        )
    else:
        # Preserve the low-level helper's historic behavior for tests/non-Unit callers.
        payload = units
    with open(pickle_path, "wb") as f:
        pickle.dump(payload, f, protocol=pickle.HIGHEST_PROTOCOL)
    print(f"Loader: Pickle cache saved to {pickle_path} ({len(units)} units)")


def _save_csv_cache(excel_path: str, units: list[Unit]) -> None:
    """Write the current unit list to CSV cache (legacy format)."""
    csv_path = _csv_cache_path(excel_path)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=_CSV_FIELDS)
        writer.writeheader()
        for unit in units:
            writer.writerow(_unit_to_csv_row(unit))
    print(f"Loader: CSV cache saved to {csv_path} ({len(units)} units)")


def save_csv_cache(excel_path: str, units: list[Unit]) -> None:
    """Public wrapper — update both caches after an in-memory change."""
    _save_pickle_cache(excel_path, units)
    _save_csv_cache(excel_path, units)


def _cache_is_fresh(excel_path: str) -> bool:
    """Return True if the Pickle cache exists and is newer than the Excel file."""
    pickle_path = _cache_path(excel_path)
    if not os.path.exists(pickle_path):
        return False
    try:
        cache_mtime = os.path.getmtime(pickle_path)
        excel_mtime = os.path.getmtime(excel_path)
        if cache_mtime < excel_mtime:
            return False
        with open(pickle_path, "rb") as f:
            payload = WorkbookCache.from_pickle(pickle.load(f))
        if payload.excel_mtime_ns and payload.excel_size:
            mtime_ns, size = _workbook_signature(excel_path)
            return payload.excel_mtime_ns == mtime_ns and payload.excel_size == size
        return True
    except (OSError, pickle.PickleError, EOFError, AttributeError):
        return False


def _csv_cache_is_fresh(excel_path: str) -> bool:
    """Return True if the legacy CSV cache can satisfy a cache-first load."""
    csv_path = _csv_cache_path(excel_path)
    if not os.path.exists(csv_path):
        return False
    try:
        return os.path.getmtime(csv_path) >= os.path.getmtime(excel_path)
    except OSError:
        return False


def _load_any_cache(excel_path: str, detailer_schedules: dict) -> list[Unit] | None:
    """Load any available cache as a rescue path when Excel is unreadable."""
    pickle_path = _cache_path(excel_path)
    if os.path.exists(pickle_path):
        print(f"Loader: Excel unreadable; rescuing from Pickle cache {pickle_path}")
        return _load_units_from_pickle(pickle_path, detailer_schedules)

    csv_path = _csv_cache_path(excel_path)
    if os.path.exists(csv_path):
        print(f"Loader: Excel unreadable; rescuing from CSV cache {csv_path}")
        units = _load_units_from_csv(csv_path, detailer_schedules)
        _save_pickle_cache(excel_path, units)
        return units

    return None


# ── Main loader ──────────────────────────────────────────────────


def load_units(
    excel_path: str,
    sheet_name: str = "Sheet1",
    detailer_schedules: dict | None = None,
    force_reload: bool = False,
) -> list[Unit]:
    if detailer_schedules is None:
        detailer_schedules = {}

    # Fast path: use Pickle cache if it's fresh and we're not forcing a reload.
    if not force_reload and (_cache_is_fresh(excel_path) or _csv_cache_is_fresh(excel_path)):
        pickle_path = _cache_path(excel_path)
        csv_path = _csv_cache_path(excel_path)

        # Prefer Pickle cache (faster), fall back to CSV cache
        if os.path.exists(pickle_path) and _cache_is_fresh(excel_path):
            print(f"Loader: Loading from Pickle cache {pickle_path}")
            start_time = time.time()
            units = _load_units_from_pickle(pickle_path, detailer_schedules)
            elapsed = time.time() - start_time
            print(f"Loader: Loaded {len(units)} units from Pickle cache in {elapsed:.2f}s.")
            return units
        if os.path.exists(csv_path) and _csv_cache_is_fresh(excel_path):
            print(f"Loader: Loading from CSV cache {csv_path}")
            start_time = time.time()
            units = _load_units_from_csv(csv_path, detailer_schedules)
            elapsed = time.time() - start_time
            print(f"Loader: Loaded {len(units)} units from CSV cache in {elapsed:.2f}s.")
            # Create Pickle cache for next time
            _save_pickle_cache(excel_path, units)
            return units

    # Slow path: parse the Excel workbook.
    default_schedule = detailer_schedules.get("default", [0, 1, 2, 3])
    print(f"Loader: Loading workbook from {excel_path} (sheet: {sheet_name})")
    start_time = time.time()
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True, keep_vba=False)
        ws = wb[sheet_name]
    except (BadZipFile, InvalidFileException, OSError) as exc:
        cached_units = _load_any_cache(excel_path, detailer_schedules)
        if cached_units is not None:
            print(f"Loader: Using cache because Excel could not be read: {exc}")
            return cached_units
        raise
    units = []

    today = datetime.now().date()
    three_months_back = today - timedelta(days=90)
    twelve_months_forward = today + timedelta(days=365)

    print(f"Loader: Filtering units from {three_months_back} to {twelve_months_forward}")

    col_to_field: dict[int, str] = {
        ord(col.upper()) - ord("A") + 1: field for field, col in COLUMN_MAP.items()
    }
    max_col = max(col_to_field)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=max_col, values_only=True), start=2):
        row_data = {
            field: row[col_idx - 1]
            for col_idx, field in col_to_field.items()
            if col_idx - 1 < len(row)
        }

        detailer_name = str(row_data.get("detailer", ""))
        schedule = detailer_schedules.get(detailer_name, default_schedule)

        unit = Unit(
            com_number=str(row_data.get("com_number", "") or f"ROW{row_idx}"),
            job_name=str(row_data.get("job_name", "")),
            contract_number=str(row_data.get("contract_number", "")),
            description=str(row_data.get("description", "")),
            detailer=detailer_name,
            checking_status=str(row_data.get("checking_status", "")),
            department_hours=parse_float(row_data.get("department_hours")),
            working_days=schedule,
            target_department_hours=parse_float(row_data.get("target_department_hours")),
            iec_internal_hours=parse_float(row_data.get("iec_internal_hours")),
            percent_complete=parse_float(row_data.get("percent_complete")) * 100,
            unit_detailing_start_date=parse_date(row_data.get("unit_detailing_start_date")),
            unit_moved_to_checking_date=parse_date(row_data.get("unit_moved_to_checking_date")),
            unit_detailing_completion_date=parse_date(
                row_data.get("unit_detailing_completion_date")
            ),
            dept_due_date_previous=parse_date(row_data.get("dept_due_date_previous")),
            detailing_due_date=parse_date(row_data.get("detailing_due_date")),
            build_date=parse_date(row_data.get("build_date")),
        )

        # Filter logic
        should_include = False
        for _, d in unit.milestones:
            if d and three_months_back <= d <= twelve_months_forward:
                should_include = True
                break

        if should_include:
            unit.status_color = unit.calculated_status_color
            unit.excel_row = row_idx
            unit.fingerprint = unit_fingerprint(unit)
            units.append(unit)

    wb.close()
    elapsed = time.time() - start_time
    print(
        f"Loader: Finished loading. {len(units)} units included after filtering in {elapsed:.2f}s."
    )

    # Save both caches for next time.
    _save_pickle_cache(excel_path, units)
    _save_csv_cache(excel_path, units)

    return units
