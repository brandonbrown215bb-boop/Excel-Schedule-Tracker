# data/writer.py
from __future__ import annotations

import gc
import os
import shutil
import tempfile
from contextlib import suppress
from zipfile import is_zipfile, ZipFile

from openpyxl import load_workbook  # type: ignore[reportMissingImports]
from openpyxl.utils.cell import column_index_from_string  # type: ignore[reportMissingImports]
from openpyxl.workbook.workbook import Workbook  # type: ignore[reportMissingImports]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[reportMissingImports]

from data.loader import COLUMN_MAP
from data.models import Unit

# Reverse map: column letter → field name
COL_TO_FIELD = {v: k for k, v in COLUMN_MAP.items()}

# Removed: STATUS_FILLS = { ... }


def _safe_save_workbook(wb: Workbook, excel_path: str) -> None:
    """Save to a temp workbook first, then replace the target atomically."""
    directory = os.path.dirname(os.path.abspath(excel_path)) or "."
    suffix = os.path.splitext(excel_path)[1] or ".xlsx"
    fd, temp_path = tempfile.mkstemp(prefix=".unittracker-save-", suffix=suffix, dir=directory)
    os.close(fd)

    backup_path = excel_path + ".bak"
    try:
        wb.save(temp_path)

        # Validate temp file without leaking ZipFile handles.
        # openpyxl's save() can leave internal ZipFile references that, when
        # GC'd after os.replace(), trigger "I/O operation on closed file".
        if not _is_zipfile_safe(temp_path):
            raise ValueError("Temporary workbook save did not produce a valid Excel zip file")

        if os.path.exists(excel_path) and _is_zipfile_safe(excel_path):
            shutil.copy2(excel_path, backup_path)

        # Close any internal ZipFile handles openpyxl may have left dangling.
        # Not part of openpyxl's public API, but prevents noisy GC warnings
        # on CPython 3.14+ when the temp file is os.replace'd away.
        for attr in ("_archive", "_zip"):
            handle = getattr(wb, attr, None)
            if handle is not None:
                with suppress(Exception):
                    handle.close()

        # Force GC to collect any remaining ZipFile references before replacing.
        gc.collect()
        os.replace(temp_path, excel_path)
    finally:
        if os.path.exists(temp_path):
            with suppress(OSError):
                os.remove(temp_path)


def _is_zipfile_safe(path: str) -> bool:
    """Check if *path* is a valid ZIP without leaking file handles.

    :func:`zipfile.is_zipfile` can leave the underlying file handle open
    in some Python versions, causing ``ValueError: I/O operation on
    closed file`` when the ZipFile is later garbage-collected after the
    file has been moved.
    """
    try:
        with ZipFile(path, "r") as _:
            return True
    except (OSError, ValueError):
        return False


def find_row_by_com(ws: Worksheet, com_number: str) -> int | None:
    """Find the row index for a given COM number."""
    com_col = column_index_from_string(COLUMN_MAP["com_number"])
    for row_idx in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=com_col)
        if str(cell.value).strip() == str(com_number).strip():
            return row_idx
    return None


def _get_worksheet(wb: Workbook, sheet_name: str) -> Worksheet:
    """Return a normal worksheet from a workbook, rejecting chart/write-only sheets."""
    sheet = wb[sheet_name]
    if not isinstance(sheet, Worksheet):
        raise TypeError(f"Sheet {sheet_name!r} is not a writable worksheet")
    return sheet


def _unit_field_values(unit: Unit) -> dict[str, object]:
    """Map Unit fields to Excel-ready values."""
    return {
        "com_number": unit.com_number,
        "job_name": unit.job_name,
        "contract_number": unit.contract_number,
        "description": unit.description,
        "detailer": unit.detailer,
        "checking_status": unit.checking_status,
        "department_hours": unit.department_hours,
        "actual_hours": unit.actual_hours,
        "target_department_hours": unit.target_department_hours,
        "iec_internal_hours": unit.iec_internal_hours,
        "percent_complete": unit.percent_complete / 100.0,
        "unit_detailing_start_date": unit.unit_detailing_start_date,
        "unit_moved_to_checking_date": unit.unit_moved_to_checking_date,
        "unit_detailing_completion_date": unit.unit_detailing_completion_date,
        "dept_due_date_previous": unit.dept_due_date_previous,
        "detailing_due_date": unit.detailing_due_date,
        "build_date": unit.build_date,
    }


def _update_bak(excel_path: str) -> None:
    """Create/update the .bak backup after a fast-path save."""
    backup_path = excel_path + ".bak"
    try:
        if os.path.exists(excel_path):
            shutil.copy2(excel_path, backup_path)
    except OSError:
        pass  # non-critical; skip


def save_unit(
    excel_path: str,
    unit: Unit,
    sheet_name: str = "Sheet1",
    row_idx: int | None = None,
    force: bool = False,
) -> None:
    """Write a unit\'s data back to its row in the Excel file.

    Attempts the ZIP/XML fast path first; falls back to full openpyxl
    load-modify-save if the fast path cannot proceed.

    Args:
        excel_path: Path to the .xlsm/.xlsx workbook.
        unit: The unit to write.
        sheet_name: Worksheet name within the workbook.
        row_idx: Cached row hint (skips COM column scan).
        force: If True, skip the COM-number row validation and write
            unconditionally to the indicated row. Used for conflict
            overwrite after a :class:`RevisionConflictError`.
    """
    # Fast path: direct ZIP/XML modification
    try:
        from data.fast_writer import save_unit_fast

        # Map "Sheet1" (default) to "Current List" (actual sheet name)
        fast_sheet = "Current List" if sheet_name in ("Sheet1", "Current List") else sheet_name
        if save_unit_fast(excel_path, unit, fast_sheet):
            _update_bak(excel_path)
            return
    except ImportError:
        pass  # fast_writer module not available — use standard path

    # Fallback: existing openpyxl path
    wb = load_workbook(excel_path, keep_vba=True)
    try:
        ws = _get_worksheet(wb, sheet_name)

        if row_idx is None:
            row_idx = unit.excel_row
        if row_idx is not None and not force:
            com_col = column_index_from_string(COLUMN_MAP["com_number"])
            existing_com = ws.cell(row=row_idx, column=com_col).value
            if str(existing_com).strip() != str(unit.com_number).strip():
                row_idx = None
        if row_idx is None:
            row_idx = find_row_by_com(ws, unit.com_number)
        if row_idx is None:
            raise ValueError(f"COM number {unit.com_number} not found in sheet")

        for field_name, value in _unit_field_values(unit).items():
            col = COLUMN_MAP.get(field_name)
            if col:
                cell = ws.cell(row=row_idx, column=column_index_from_string(col))
                if isinstance(value, float) and field_name == "percent_complete":
                    cell.value = value
                    cell.number_format = "0%"
                else:
                    cell.value = value

        # Removed: Update status color on the Detailing Due Date cell (column A)
        # The status color is now derived in data/models.py and not saved to Excel directly.
        # If Excel needs conditional formatting, it should be set up in Excel itself.

        _safe_save_workbook(wb, excel_path)
    finally:
        wb.close()
